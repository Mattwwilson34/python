from openpyxl import load_workbook
from pprint import pprint

workbook = load_workbook(filename='ATM.xlsx')
workbook.sheetnames

sheet_visit_1 = workbook[workbook.sheetnames[0]]
sheet_visit_2_6 = workbook[workbook.sheetnames[1]]

sheet_1_titles = []
sheet_2_titles = []

# Get all title from first sheet_visit_1
for column in sheet_visit_1.iter_cols(min_row=1, max_row=1, min_col=1, max_col=50):
    if column[0].value is None:
        continue
    else:
        sheet_1_titles.append(column[0].value)


for column in sheet_visit_2_6.iter_cols(min_row=1, max_row=1, min_col=1, max_col=50):
    if column[0].value is None:
        continue
    else:
        sheet_2_titles.append(column[0].value)


def print_row_cells(sheet, min, max):
    for row in sheet.iter_rows(min_row=min, max_row=max, values_only=True):
        print(row)


def get_row_cells(sheet, row_number):
    arr_of_cells = []
    for row in sheet.iter_rows(min_row=row_number, max_row=row_number):
        arr_of_cells.append(row)
    return arr_of_cells


def get_column_cells(sheet, column_number):
    arr_of_cells = []
    for cells in sheet.iter_cols(min_col=column_number, max_col=column_number):
        for cell in cells:
            arr_of_cells.append(cell.value)
    return arr_of_cells


def get_column_depth(sheet):
    for col_cells in sheet.iter_cols(min_row=1, min_col=1, max_col=1):
        return len(col_cells)


l1 = get_row_cells(sheet_visit_1, 1)[0]
l2 = get_row_cells(sheet_visit_2_6, 1)[0]

l1Dict = {}

for cell in l1:
    l1Dict[cell.value] = cell.column

l2Dict = {}

for cell in l2:
    l2Dict[cell.value] = cell.column


def append_data(sheet, data, col):
    start = 14
    for cell in data:
        sheet.cell(row=start, column=col).value = cell
        start += 1


pprint(l2Dict)

for cell2 in l2:
    for cell1 in l1:
        if cell1.value == cell2.value:
            col_appending_to = l1Dict[cell1.value]
            col_of_data_appending = l2Dict[cell1.value]
            data = get_column_cells(sheet_visit_2_6, col_of_data_appending)
            append_data(sheet_visit_1, data, col_appending_to)

        else:
            continue

workbook.save(filename='ATM.xlsx')
