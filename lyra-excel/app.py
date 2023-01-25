from openpyxl import load_workbook

import sys
import utils

# ROW_TO_BEGIN_INSERTION = sys.argv[2]
# SPREADSHEET_FILENAME = sys.argv[1]


def app():
    workbook = load_workbook(filename='ATM.xlsx')
    workbook.sheetnames

    sheet_1 = workbook[workbook.sheetnames[0]]
    sheet_2 = workbook[workbook.sheetnames[1]]

    l1 = utils.get_row(sheet_1, 1)
    l2 = utils.get_row(sheet_2, 1)

    l1Dict = utils.build_dictonary_from_cell_array(l1)
    l2Dict = utils.build_dictonary_from_cell_array(l2)

    # iterate over the smaller list of columns title cells
    for cell2 in l2:
        # iterate over the larger list of colum title cells
        for cell1 in l1:
            # if the titles of the 2 title cells match append data from sheet 2
            if cell1.value == cell2.value:
                # get sheet 1 column number from dictionary
                col_appending_to = l1Dict[cell1.value]
                # get sheet 2 column number from dictionary
                col_appending_from = l2Dict[cell1.value]
                # get data cells array from sheet 2
                data = utils.get_column_cells(
                    sheet_2, col_appending_from)
                # append data to sheet 1
                utils.append_data_to_sheet(
                    sheet_1, data, col_appending_to, 14)

            # if there is not a match of titles then do nothing
            else:
                continue

    workbook.save(filename='ATM.xlsx')


app()
